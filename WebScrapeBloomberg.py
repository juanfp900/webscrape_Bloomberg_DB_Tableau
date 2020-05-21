import requests
import re
import math
from bs4 import BeautifulSoup as bs
import pandas as pd
from sqlalchemy import create_engine 
from openpyxl import load_workbook
import xlsxwriter
import pycountry
import ccy
import currency
import pycountry
from currency_converter import CurrencyConverter
from forex_python.converter import CurrencyRates
from pycountry_convert import country_name_to_country_alpha2


class WebScrapeBloomberg:

	def RequestBloombergData(self, link):
		headers = {
		    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
		    'referrer': 'https://google.com',
		    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
		    'Accept-Encoding': 'gzip, deflate, br',
		    'Accept-Language': 'en-US,en;q=0.9',
		    'Pragma': 'no-cache'}

		webRequest = requests.get(link, headers=headers, verify = False)

		data = webRequest.content
		rawData = bs(data, 'html.parser')
		return(rawData)


	def FindCountryHeaders(self, rawData):
		countryList = []
		#Parse to get country name
		countryHeaders = rawData.find_all("div", {"class" : "table-container__header"})
		#loop through nexted classes containing country Name
		for header in countryHeaders:
			temp = str(header.find_all("h2", {"class" : "table-container__title"}))
			temp2 = temp.split(">")[1]
			temp3 = temp2.split("<")[0]
			if temp3 == "ID":
				countryList.append("Indonesia")
			else:
				countryList.append(temp3)
		return(countryList)


	#Name, Net_change &Change, 1Month, 1Year, Time(EDT)
	#Only need to do this once. Save results in list
	def FindTableHeaders(self,rawData):
		tableHeaderList = []
		tableHeaders = rawData.find_all("th", {"class" : "data-table-headers-cell"})
		counter = 0
		for header in tableHeaders:
			if counter == 7:
				break;
			else:
				header = str(header)	#TypeError: 'NoneType' object is not callable. So convert to str
				temp2 = header.split(">")[1]
				temp3 = temp2.split("<")[0]
				temp3 = str(temp3.lower())
				tableHeaderList.append(temp3)
				counter = counter + 1
		return(tableHeaderList)


	#indexNames
	def FindRowName(self,rawData):
		indexNames = rawData.find_all("div", {"class" : "data-table-row-cell__link-block"})
		indexNamesList = []
		counter = 0
		for index in indexNames:
			index = str(index)
			temp2 = index.split(">")[1]
			temp3 = temp2.split("<")[0]
			if counter == 0 or counter % 2 == 0:
				counter = counter + 1 
			else:
				indexNamesList.append(temp3)
				counter = counter + 1
		indexNamesList = [x.strip(' ') for x in indexNamesList] #remove whitespace
		
		return(indexNamesList)


	def CombineLists(self,rawData,indexNamesList):
		combinedList = []
		list = self.ConvertRawToList(rawData)	#only called once in the first loop

		for indexName in indexNamesList:
			rowList, list = self.ParseList(list, indexName)
			combinedList.append(rowList)
		return(combinedList)
		
		

	def ConvertRawToList(self,rawData):
		rowValues = rawData.find_all("span", {"class" : "data-table-row-cell__value"})
		list = []
		
		for row in rowValues:
			row = str(row)
			temp2 = row.split(">")[1]
			temp3 = temp2.split("<")[0]
			list.append(temp3)
		list.append("empty")	#append random value list to deal with empty value condition
		return(list)


	def ParseList(self, list, indexName):
		rowList = []
		rowList.append(indexName)
		count = 0
		for value in list:
			value = str(value)
			if count == 6:
				list = list[6:]	#remove 6 first values in the list
				#print("current list")
				#print(list)
				return (rowList, list)
			else:
				rowList.append(value)
				count = count + 1



	#assigns a country to each row we got from combineLists method
	
	def AssignAmericasToList(self, countryList, combinedList):
			newList = [] #Create new list to assign ll values
			count = 0
			for i in combinedList:
				if count <= 21:
					i.append(countryList[0]) #append United States
					newList.append(i) #append list to newList
				if count == 22 or count == 23:
					i.append(countryList[1]) #append Argentina
					newList.append(i) 
				if count == 24 or count == 25:
					i.append(countryList[2]) #append Peru
					newList.append(i) 
				if count == 26 or count == 27:
					i.append(countryList[3]) #append Brazil
					newList.append(i) 
				if count == 28 or count == 29:
					i.append(countryList[4]) #append Mexico
					newList.append(i) 
				if count == 30 or count == 31:
					i.append(countryList[5]) #append Canada
					newList.append(i) 
				if count == 32 or count == 33:
					i.append(countryList[6]) #append Chile
					newList.append(i) 
				if count == 34:
					i.append(countryList[7]) #append Venezuela
					newList.append(i) 
				if count == 35:
					i.append(countryList[8]) #append Costa Rica
					newList.append(i) 
				if count == 36:
					i.append(countryList[9]) #append Panama
					newList.append(i) 
				if count == 37:
					i.append(countryList[10]) #append Jamaica
					newList.append(i) 
				if count == 38:
					i.append(countryList[11]) #append Colombia
					newList.append(i) 
				if count == 39:
					i.append(countryList[12]) #append Bermuda
					newList.append(i) 
				count = count + 1
			return(newList)

	
	def AssignEMEAToList(self, countryList, combinedList):
		#print(countryList)
		newList = [] #Create new list to assign ll values
		count = 0
		for i in combinedList:
			if count <= 3:
				i.append(countryList[0]) #append Europe
				newList.append(i) 
			if count == 4 or count == 5 or count == 6:
				i.append(countryList[1]) #append United Kingdom
				newList.append(i) 
			if count == 7 or count == 8 or count == 9:
				i.append(countryList[2]) #append Denmark
				newList.append(i)
			if count == 10 or count == 11 or count == 12:
				i.append(countryList[3]) #append Poland
				newList.append(i)
			if count == 13 or count == 14:
				i.append(countryList[4]) #append France
				newList.append(i) 
			if count == 15 or count == 16:
				i.append(countryList[5]) #append Turkey ##UP UNTIL HERE
				newList.append(i)
			if count == 17 or count == 18:
				i.append(countryList[6]) #append Bahrain
				newList.append(i)
			if count == 19 or count == 20:
				i.append(countryList[7]) #append Germany
				newList.append(i)
			if count == 21 or count == 22:
				i.append(countryList[8]) #append United Arab Emirates
				newList.append(i)
			if count == 23 or count == 24:
				i.append(countryList[9]) #append Russia
				newList.append(i)
			if count == 25 or count == 26:
				i.append(countryList[10]) #append Greece
				newList.append(i)
			if count == 27 or count == 28:
				i.append(countryList[11]) #append South Africa
				newList.append(i)  
			if count == 29 or count == 30:
				i.append(countryList[12]) #append Finland
				newList.append(i)
			if count == 31 or count == 32:
				i.append(countryList[13]) #append Kuwait
				newList.append(i)
			if count == 33 or count == 34:
				i.append(countryList[14]) #append Ukraine
				newList.append(i) 
			if count == 35 or count == 36:
				i.append(countryList[15]) #append Switzerland
				newList.append(i) 
			if count == 37 or count == 38:
				i.append(countryList[16]) #append Kenya
				newList.append(i)
			if count == 39 or count == 40:
				i.append(countryList[17]) #append Serbia
				newList.append(i)
			if count == 41 or count == 42:
				i.append(countryList[18]) #append Norway
				newList.append(i) 
			if count == 43 or count == 44:
				i.append(countryList[19]) #append Hungary
				newList.append(i)
			if count == 45 or count == 46:
				i.append(countryList[20]) #append Spain
				newList.append(i) 
			if count == 47 or count == 48:
				i.append(countryList[21]) #append Netherlands
				newList.append(i)
			if count == 49 or count == 50:
				i.append(countryList[22]) #append Italy
				newList.append(i)
			if count == 51 or count == 52:
				i.append(countryList[23]) #append Bosnia and Herzegovina
				newList.append(i)
			if count == 53 or count == 54:
				i.append(countryList[24]) #append Belgium
				newList.append(i)
			if count == 55 or count == 56:
				i.append(countryList[25]) #append Luxembourg
				newList.append(i)
			if count == 57 or count == 58:
				i.append(countryList[26]) #append Portugal
				newList.append(i)
			if count == 59 or count == 60:
				i.append(countryList[27]) #append Israel
				newList.append(i) 
			if count == 61 or count == 62:
				i.append(countryList[28]) #append Macedonia
				newList.append(i)
			if count == 63 or count == 64:
				i.append(countryList[29]) #append Austria
				newList.append(i)
			if count == 65 or count == 66:
				i.append(countryList[30]) #append Sweden
				newList.append(i) 
			if count == 67:
				i.append(countryList[31]) #append Tunisia
				newList.append(i)
			if count == 68:
				i.append(countryList[32]) #append Montenegro
				newList.append(i)   
			if count == 69:
				i.append(countryList[33]) #append Saudi Arabia
				newList.append(i) 
			if count == 70:
				i.append(countryList[34]) #append Botswana
				newList.append(i) 
			if count == 71:
				i.append(countryList[35]) #append Qatar
				newList.append(i) 
			if count == 72:
				i.append(countryList[36]) #append Jordan
				newList.append(i)
			if count == 73:
				i.append(countryList[37]) #append Slovenia
				newList.append(i)
			if count == 74:
				i.append(countryList[38]) #append Palestine
				newList.append(i)
			if count == 75:
				i.append(countryList[39]) #append Palestine
				newList.append(i)  
			if count == 76:
				i.append(countryList[40]) #append Nigeria
				newList.append(i)
			if count == 77:
				i.append(countryList[41]) #append Iceland
				newList.append(i)
			if count == 78:
				i.append(countryList[42]) #append Namibia
				newList.append(i) 
			if count == 79:
				i.append(countryList[43]) #append Cyprus
				newList.append(i) 
			if count == 80:
				i.append(countryList[44]) #append Lebanon
				newList.append(i) 
			if count == 81:
				i.append(countryList[45]) #append Oman
				newList.append(i) 
			if count == 82:
				i.append(countryList[46]) #append Mauritius
				newList.append(i) 
			if count == 83:
				i.append(countryList[47]) #append Ireland
				newList.append(i)
			if count == 84:
				i.append(countryList[48]) #append Slovakia
				newList.append(i)
			if count == 85:
				i.append(countryList[49]) #append Croatia
				newList.append(i) 
			if count == 86:
				i.append(countryList[50]) #append Morocco
				newList.append(i) 
			if count == 87:
				i.append(countryList[51]) #append Tanzania
				newList.append(i)
			if count == 88:
				i.append(countryList[52]) #append Bulgaria
				newList.append(i)
			if count == 89:
				i.append(countryList[53]) #append Malta
				newList.append(i)
			if count == 90:
				i.append(countryList[54]) #append Latvia
				newList.append(i)
			if count == 91:
				i.append(countryList[55]) #append Estonia
				newList.append(i)
			if count == 92:
				i.append(countryList[56]) #append Africa/Middle East
				newList.append(i)
			if count == 93:
				i.append(countryList[57]) #append Kazakhstan
				newList.append(i)
			if count == 94:
				i.append(countryList[58]) #append Romania
				newList.append(i)
			if count == 95:
				i.append(countryList[59]) #append Lithuania
				newList.append(i)
			if count == 96:
				i.append(countryList[60]) #append Ghana
				newList.append(i)
			if count == 97:
				i.append(countryList[61]) #append Czech Republic
				newList.append(i)

			count = count + 1
		return(newList)
			


	def AssignAsiaPacificToList(self, countryList, combinedList):
		#print(countryList)
		newList = [] #Create new list to assign ll values
		count = 0
		for i in combinedList:
			if count <= 14:
				i.append(countryList[0]) #append Japan
				newList.append(i)
			if  15 <= count <= 26:
				i.append(countryList[1]) #append China
				newList.append(i)
			if  27 <= count <= 37:
				i.append(countryList[2]) #append Hong Kong
				newList.append(i)
			if  38 <= count <= 44:
				i.append(countryList[3]) #append South Korea
				newList.append(i)
			if  45 <= count <= 49:
				i.append(countryList[4]) #append India
				newList.append(i)
			if  50 <= count <= 53:
				i.append(countryList[5]) #append New Zealand
				newList.append(i)
			if  54 <= count <= 56:
				i.append(countryList[6]) #append Taiwan
				newList.append(i)
			if  57 <= count <= 60:
				i.append(countryList[7]) #append Austalia
				newList.append(i)
			if  61 <= count <= 62:
				i.append(countryList[8]) #append Pakistan
				newList.append(i)
			if  63 <= count <= 64:
				i.append(countryList[9]) #append Malaysia
				newList.append(i)
			if  65 <= count <= 66:
				i.append(countryList[10]) #append ID
				newList.append(i)
			if  67 <= count <= 68:
				i.append(countryList[11]) #append Singapore
				newList.append(i)
			if  69 <= count <= 70:
				i.append(countryList[12]) #append Thailand
				newList.append(i)
			if  71 <= count <= 72:
				i.append(countryList[13]) #append Vietnam
				newList.append(i)
			if  count == 73:
				i.append(countryList[14]) #append Bangladesh
				newList.append(i)
			if  count == 74:
				i.append(countryList[15]) #append Mongolia
				newList.append(i)
			if  count == 75:
				i.append(countryList[16]) #append Asia Region
				newList.append(i)
			if  count == 76:
				i.append(countryList[17]) #append Laos
				newList.append(i)
			if  count == 77:
				i.append(countryList[18]) #append Philippines
				newList.append(i)
			if  count == 78:
				i.append(countryList[19]) #append SriLanka
				newList.append(i)

			count = count + 1
		return(newList)
		


	def ConvertListToDataframe(self, tableHeaders, newList):

		tableHeaders.append("country")
		df = pd.DataFrame(newList)
		df.columns = tableHeaders
		
		nameList = [] #Create new name list that removes unwanted characters

		for i in df['name']:
			i = re.sub('amp;', '', i)
			nameList.append(i)
		
		df['name'] = nameList
		df.rename(columns={'% change': 'percent change'},inplace = True)
		df.rename(columns={'time (edt)': 'eastern time'},inplace = True)

		return(df)



	def FindCountryCodes(self, df):
	
		countryCodeList = []
		df['country'] = df['country'].str.lstrip()
		for country in df['country']:
			if country == "Europe":
				countryCodeList.append("FR")	#May not use for actual analysis
				continue
			if country == "Asia Region":
				countryCodeList.append("JP")	#May not use for actual analysis
				continue
			if country == "Africa/Middle East":	#May not use for actual analysis
				countryCodeList.append("AE")	#UAE
				continue
			else:
				val = country_name_to_country_alpha2(country)
				countryCodeList.append(val)

		df['country_code'] = countryCodeList
		#print(df)
		#print(countryCodeList)

		return(df)
		


	def FindCurrencyCodes(self, df, c):

		currencyList = []
		for countryCode in df["country_code"]:
			countryCode = str(countryCode)
			c = ccy.countryccy(countryCode)
			c = str(c)
			
			if c != "None":
				currencyList.append(c)
			elif c == "None":
				if countryCode == "VE":		#Venezuela
					currencyList.append("VEF")
				
				if countryCode == "CR":		#Costa Rica
					currencyList.append("CRC")
			
				if countryCode == "PA":		#Panama
					currencyList.append("PAB")
				
				if countryCode == "BH":		#Baharin //Ccheck
					currencyList.append("BHD")

				if countryCode == "KW":		#Kuwait	//check
					currencyList.append("KWD")
				
				if countryCode == "KE":		#Kenya	//check
					currencyList.append("KES")

				if countryCode == "RS":		#Serbia	//check
					currencyList.append("RSD")

				if countryCode == "ME":		#Montenegro	//check
					currencyList.append("EUR")			
				
				if countryCode == "MK":		#Macedonia	//check
					currencyList.append("MKD")
				
				if countryCode == "BA":		#Bosnia and Herzegovina
					currencyList.append("BAM")
				
				if countryCode == "MU":		#Mauritius	//check
					currencyList.append("MUR")

				if countryCode == "MA":		#Moroccan	//check
					currencyList.append("MAD")

				if countryCode == "PS":		#Palestine	//check
					currencyList.append("ILS")

				if countryCode == "TN":		#Tunisia	//check
					currencyList.append("TND")

				if countryCode == "BW":		#Botswana //check
					currencyList.append("BWP")
				
				if countryCode == "JO":		#Jordan		//check
					currencyList.append("JOD")

				if countryCode == "IS":		#Iceland	//check
					currencyList.append("ISK")

				if countryCode == "NA":		#Namibia //check
					currencyList.append("NAD")

				if countryCode == "LB":		#Lebanon //check
					currencyList.append("LBP")

				if countryCode == "OM":		#Oman	//check
					currencyList.append("OMR")

				if countryCode == "TZ":		#Tanzania	//check
					currencyList.append("TZS")

				if countryCode == "GH":		#Ghana	//check
					currencyList.append("GHS")

				if countryCode == "UA":	#Ukraine	//check
					currencyList.append("UAH")

				if countryCode == "PK":	#Pakistan	
					currencyList.append("PKR")

				if countryCode == "BD":	#Bangladesh	
					currencyList.append("BDT")

				if countryCode == "MN":	#Mongolia
					currencyList.append("MNT")

				if countryCode == "LA":	#Laos
					val = str("LAK")
					currencyList.append(val)

				if countryCode == "LK":	#Sri Lanka
					val = str("LKR")
					currencyList.append(val)
			

		df['countryCurrency'] = currencyList
		#pd.set_option('display.max_rows', None)
		#pd.set_option('display.max_column', None)
		return(df)

	

	def FindCurrencySymbols(self,df):
		symbolsList = []
		for countryCurrency in df["countryCurrency"]:
			countryCurrency = str(countryCurrency)
			if countryCurrency == 'BMD':
				symbolsList.append("$")
				continue
			if countryCurrency == "MNT":
				symbolsList.append("₮")
				continue
			if countryCurrency == "LAK":
				symbolsList.append("₭")
				continue
			else:
				symbol = currency.symbol(countryCurrency)
				symbol = str(symbol)
				symbolsList.append(symbol)
			
	
		df['currency_symbol'] = symbolsList
		return(df)

	def FindCurrencyNames(self,df):
		currencyNameList = []
		for countryCurrency in df["countryCurrency"]:
			countryCurrency = str(countryCurrency)
			if countryCurrency == 'BMD':
				currencyNameList.append("Bermuda Dollar")
				continue
			if countryCurrency == "MNT":
				currencyNameList.append("Mongolian Tögrög")
				continue
			if countryCurrency == "LAK":
				currencyNameList.append("Laotian Kip")
				continue
			else:
				name = currency.name(countryCurrency)
				name = str(name)
				currencyNameList.append(name)
			
	
		df['currency_name'] = currencyNameList
		return(df)


	def CurrencyConversion(self, df, c):
		
		#going to need country_currency column
		usdExchangeList = []
		for countryCurrency in df["countryCurrency"]:
			countryCurrency = str(countryCurrency)
			countryCurrency = countryCurrency.strip()
			
			if countryCurrency == 'BMD':
				usdExchangeList.append(1)
				continue
			if countryCurrency == "MNT":
				usdExchangeList.append(0.000359195)
				continue
			if countryCurrency == "LAK":
				usdExchangeList.append(0.00011)	# 1 LAK = 0.00011 USD
				continue
			if countryCurrency == 'USD':
				usdExchangeList.append(1)
				continue
			if countryCurrency ==  'ARS':
				usdExchangeList.append(0.015)
				continue
			else:
				exchange = c.get_rate(countryCurrency, 'USD') #convert currencies to USD
				print(exchange)
				usdExchangeList.append(exchange)
				
			
	
		df['value_in_usd'] = usdExchangeList
		print(df)

		

	def DfToDatabase(self, df):
		
		tableNames = ['americas', 'emea', 'asia_pacific']
		#engine = create_engine('postgresql://:Nbareddit@12@Juans-MacBook-Air.local:3306/world_financial_indexes')
		engine = create_engine('DBtype://Password@localhost:port/databaseName')
		connection = engine.connect()

		currVal = df['country'].iloc[0:1]

		if df['country'].str.contains('United States').any(): #pick a cell within df['Country']
			table_name = tableNames[0]
			df.to_sql(table_name,connection, if_exists = 'replace')
			#df.to_sql(tableNames, connection, index=False, if_exists='replace')
	

		if df['country'].str.contains('Europe').any():		#pick a cell within df['Country']
			table_name = tableNames[1]
			df.to_sql(table_name, connection, index=False, if_exists='replace', 
					chunksize=25000, method=None)
			

		if df['country'].str.contains('Japan').any(): #pick a cell within df['Country']
			table_name = tableNames[2]
			df.to_sql(table_name, connection, index=False, if_exists='replace', 
					chunksize=25000, method=None)

		print(engine.table_names())
		connection.close()
		


	def DfToExcel(self, df):
		today = pd.Timestamp("today").strftime("%m/%d/%Y")
		df['date'] = today

		currentTime = pd.datetime.now().strftime("%I:%M:%S")
		df['time'] = currentTime
	

		pd.set_option('display.max_rows', None)
		pd.set_option('display.max_column', None)
		print(df)
		
		path = r'/Users/JuanP/Desktop/PythonProjects/WebScrape_Bloomberg_DB_Tableau/World_Financial_Indexes.xlsx'
		book = load_workbook(path)

		writer = pd.ExcelWriter(path, engine = 'openpyxl')
		writer.book = book

		if df['country'].str.contains('United States').any(): #pick a cell within df['Country']
			df.to_excel(writer,"Americas", index=False)
		
		if df['country'].str.contains('Europe').any():	#pick a cell within df['Country']
			df.to_excel(writer,"EMEA", index=False)
	
		if df['country'].str.contains('Japan').any(): #pick a cell within df['Country']
			df.to_excel(writer,"AsiaPacific", index=False)
		writer.save()
		print("Sheet Saved.")



	def DfToExcelPre(self, df, val):
		today = pd.Timestamp("today").strftime("%m/%d/%Y")
		df['date'] = today

		currentTime = pd.datetime.now().strftime("%I:%M:%S")
		df['time'] = currentTime
		

		path = r'/Users/JuanP/Desktop/PythonProjects/WebScrape_DB_Tableau/cmon2.xlsx'
		book = load_workbook(path)



		writer = pd.ExcelWriter(path, engine = 'openpyxl', mode = 'a')
		writer.book = book
		
		print("length of df")

		#40 = americas df length
		#97 = EMEA df length
		#79 = AsiaPacific df length
		print(len(df))
		print("counter")
		print(val)

		if df['country'].str.contains('United States').any(): #pick a cell within df['Country']
			if val == 1:
				df.to_excel(writer, "Americas", index=False)
			else:
				startrow = len(df) * val + 2
				df.to_excel(writer, "Americas", startrow, index=False)

		if df['country'].str.contains('Europe').any():	#pick a cell within df['Country']
			if val == 1:
				df.to_excel(writer,"EMEA", index=False)
			if 'EMEA' in book.sheetnames:
				startrow = len(df) * val + 2
				df.to_excel(writer,"EMEA", startrow, index=False)
	
		if df['country'].str.contains('Japan').any(): #pick a cell within df['Country']
			if val == 1:
				df.to_excel(writer,"AsiaPacific", index=False)
			if 'AsiaPacific' in book.sheetnames:
				startrow = len(df) * val + 2
				df.to_excel(writer,"AsiaPacific", startrow, index=False)
	
		writer.save()
		print("Sheet Saved.")


	# This method will read a counter everytime the script is ran
	def Get_var_value(self):
		filename = "scriptCounter4.dat"
		with open(filename, "a+") as f:
			f.seek(0)
			val = int(f.read() or 0) + 1
			f.seek(0)
			f.truncate()
			f.write(str(val))
			return val



	def BloombergLinks(self, obj, val, c):
		bloombergLinks = ['https://www.bloomberg.com/markets/stocks/world-indexes/americas',
		'https://www.bloomberg.com/markets/stocks/world-indexes/europe-africa-middle-east',
		'https://www.bloomberg.com/markets/stocks/world-indexes/asia-pacific']

		count = 0;
		for link in bloombergLinks:
	 		rawData = obj.RequestBloombergData(link)
	 		countryList = obj.FindCountryHeaders(rawData)
	 		tableHeaders = obj.FindTableHeaders(rawData)  
	 		indexNamesList = obj.FindRowName(rawData)
	 		combinedList = obj.CombineLists(rawData,indexNamesList)
	 		if count == 0:
	 			newList = obj.AssignAmericasToList(countryList, combinedList)
	 		if count == 1:
	 			newList = obj.AssignEMEAToList(countryList, combinedList)
	 		if count == 2:
	 			newList = obj.AssignAsiaPacificToList(countryList, combinedList)
	 		df = obj.ConvertListToDataframe(tableHeaders, newList)

	 		df = obj.FindCountryCodes(df)
	 		df = obj.FindCurrencyCodes(df,c)
	 		df = obj.FindCurrencySymbols(df)
	 		df = obj.FindCurrencyNames(df)
	 		#df = obj.CurrencyConversion(df, c)
	 		
	 		#obj.DfToDatabase(df)
	 		#obj.DfToExcelPre(df)
	 		
	 		obj.DfToExcel(df)
	 	
	 		count = count + 1


def main():
	 c = CurrencyRates()
	 obj = WebScrapeBloomberg()
	 val = obj.Get_var_value()
	 obj.BloombergLinks(obj, val, c)

main()

